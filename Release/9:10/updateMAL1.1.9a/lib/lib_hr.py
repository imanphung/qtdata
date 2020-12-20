import PySimpleGUI as sg
from datetime import datetime,timedelta
import os
import pandas as pd
import re
import base64
from openpyxl import load_workbook
import imageio
import utils
from lib import lib_sys
import cv2
        
from logger import log_temponote
import traceback
def create_department():
    mal_id = utils.ENGINE.execute("select ID from {} where DEPARTMENT like '{}'"\
            .format(utils.db_department, utils.company + '-MAL')).fetchall()[0][0]
    df = pd.read_sql('select * from {}'.format(utils.db_department), utils.ENGINE)
    df.columns = [x.upper() for x in df.columns]

    for idx, dept_name in enumerate(df['DEPARTMENT']):
        if df.loc[idx, 'ID'] == 'None':
            file_metadata = {
                'name': dept_name,
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [mal_id]
            }
            dept_id = utils.drive_service.files().create(body = file_metadata, fields = 'id').execute()['id']
            df['ID'].iloc[idx] = dept_id

    df.to_sql(utils.db_department, utils.ENGINE, if_exists = 'replace', index = False)


def create_workingspace_by_email(email, alias, dept_id):
    '''
    Create workingspace '<emailID>_<Alias>'.
    Add permission to edit.
    '''
    emailID = email.split('@')[0]

    #Create folder
    file_metadata = {
        'name': emailID + "_" + alias,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [dept_id]
        }
    folder_id = utils.drive_service.files().create(body = file_metadata, fields = 'id').execute()['id']

    #Create notes
    file_document_metadata = {
        'mimeType': 'application/vnd.google-apps.document',
        'name': emailID + "_" + alias + '_notes',
        'parents': [folder_id]
    }
    utils.drive_service.files().create(body = file_document_metadata).execute()['id']

    #Create sheets
    file_spreadsheet_metadata = {
        'mimeType': 'application/vnd.google-apps.spreadsheet',
        'name': emailID + "_" + alias + '_sheets',
        'parents': [folder_id]
    }
    utils.drive_service.files().create(body = file_spreadsheet_metadata).execute()['id']
    
    #Update id in member info
    utils.ENGINE.execute("update {} set FOLDER_ID = '{}' where EMAIL like '{}'".format(utils.db_memberinfo, folder_id, email))
    
    #Add permission
    workingspace_permission = {
        'emailAddress': email,
        'type': 'user',
        'role': 'writer'
    }
    utils.drive_service.permissions().create(fileId = folder_id, body = workingspace_permission).execute()


def create_workingspace():
    #Get members info
    df_info = pd.read_sql('select * from {}'.format(utils.db_memberinfo), utils.ENGINE)
    df_info.columns = [x.upper() for x in df_info.columns]
    #Get dept id
    df_dept = pd.read_sql('select * from {}'.format(utils.db_department), utils.ENGINE)
    df_dept.columns = [x.upper() for x in df_dept.columns]
                       
    for idx, email in enumerate(df_info['EMAIL']):
        if email != 'None' and df_info.loc[idx, 'FOLDER_ID'] == 'None':
            try:
                dept_name = df_info.loc[idx, 'DEPARTMENT']
                dept_id = df_dept.loc[df_dept['DEPARTMENT'] == dept_name, 'ID'].values[0]
                dept_permission = {
                    'emailAddress': email,
                    'type': 'user',
                    'role': 'reader'
                }
                utils.drive_service.permissions().create(fileId = dept_id, body = dept_permission).execute()

                alias = df_info.loc[idx, 'ALIAS']
                create_workingspace_by_email(email, alias, dept_id)
                
            except: 
                print(email)

#http://eyana.me/list-gdrive-folders-python/
def list_spreadsheet_by_folderid(folder_id):
    results = utils.drive_service.files().list(q = "mimeType='application/vnd.google-apps.spreadsheet' and parents in '"\
        + folder_id + "' and trashed = false", fields = "nextPageToken, files(id, name)").execute()
    items = results.get('files', [])
    
    return items

    
def list_folder_by_folderid(folder_id):
    results = utils.drive_service.files().list(q = "mimeType='application/vnd.google-apps.folder' and parents in '"\
        + folder_id + "' and trashed = false", fields = "nextPageToken, files(id, name)", pageSize = 400).execute()
    items = results.get('files', [])
        
    return items
    
    
def create_window_browser(items):
    treedata = sg.TreeData()      
    for idx, value in enumerate(items):    
        treedata.Insert('', value['name'], value['name'], values = value['id'])
    
    layout = [
        [sg.Text('Select File')],
        [sg.Tree(data = treedata, headings = ['ID'], auto_size_columns = True,
            num_rows = 20, col0_width = 30, key = '-TREE-', show_expanded = False)],
        [sg.Button('Choose')]
    ]

    window = sg.Window(
        'File Browser',
        layout,
        keep_on_top = True
    )
    
#    print(window.Element('-TREE-').TreeData.tree_dict[values['-TREE-'][0]].values)
    return window

    
def read_spreadsheet(spreadsheet_id):
    if spreadsheet_id == 'None':
        path = lib_sys.get_filepath('xlsx')
        if path is None or len(path) == 0:
            return 'No file selected.'
        
        df = pd.read_excel(path, sheetname = None)
        
        return df
    else:
        sh = utils.gc.open_by_key(spreadsheet_id)
        
        return sh


def get_memberinfo(status = 'all'):
    try:
        if status in ['in', 'temp', 'pending', 'out']:
             df = pd.read_sql("select * from {} where STATUS like '{}'".format(utils.db_memberinfo, status), utils.ENGINE)
        else:
            df = pd.read_sql("select * from {}".format(utils.db_memberinfo), utils.ENGINE)
        df.columns = [x.upper() for x in df.columns]
                      
        memberinfo_id = utils.ENGINE.execute("select ID from {} where FILENAME like '{}'"\
            .format(utils.db_file, utils.company + '-Memberinfo')).fetchall()[0][0]
        sh = utils.gc.open_by_key(memberinfo_id)
        wks = sh[0]
        wks.clear()
        wks.set_dataframe(df, 'A1')
        
        return "Members' Info downloaded successfully."
        
    except:
        return "Members' Info downloaded failed!"
    
    
def push_memberinfo():
    try:
        memberinfo_id = utils.ENGINE.execute("select ID from {} where FILENAME like '{}'"\
            .format(utils.db_file, utils.company + '-Memberinfo')).fetchall()[0][0]
        sh = utils.gc.open_by_key(memberinfo_id)
        wks = sh[0]
        df = wks.get_as_df()
        df.columns = [x.upper() for x in df.columns]
                      
        df['DATE_MODIFIED'] = datetime.now()
        df['NAME_MODIFIED'] = utils.username
        df['STARTDATE'] = pd.to_datetime(df['STARTDATE'])
        df['OUTDATE'] = pd.to_datetime(df['OUTDATE'])
                 
        df.to_sql(utils.db_memberinfo, utils.ENGINE, if_exists = 'replace', index = False)    
    
        return "Members' Info uploaded successfully."
        
    except:
        return "Members' Info uploaded failed!"

#def get_dailyreport(username = utils.username, date = utils.current_date.strftime('%Y-%m-%d')):
#    level_result = utils.check_my_level(username)
#    if level_result == 0:
#        return "You are not allowed to check on this member"
#    else:
#        date = datetime.strptime(date, '%Y-%m-%d').strftime('%d-%b-%y').upper()
#        path = utils.path_output + '{}_{}/'.format(date, username)
#        if not os.path.exists(path):
#            os.makedirs(path)
#            
#        df_note = pd.read_sql("select * from {} where NAME like '{}' and DATETIME like '{}'"\
#            .format(utils.db_temponote, username, date), utils.ENGINE)
#        df_note.columns = [x.upper() for x in df_note.columns]
#        df_note.sort_values(by='DATETIME', inplace=True, ascending=True)
#        df_note = df_note[['DATETIME', 'TIME', 'NAME', 'OUTCOME', 'NEXTACT']]
#    
#        df_action = pd.read_sql("select * from {} where NAME like '{}' and DATETIME like '{}'"\
#            .format(utils.db_action, username, date), utils.ENGINE)
#        df_action.columns = [x.upper() for x in df_action.columns]
#        df_action.sort_values(by='DATETIME', inplace=True, ascending=True)
#        df_action['NUMLIKE'] = df_action['THUMBSUP'].str.split(', ').apply(len)
#        df_action = df_action[['DATETIME', 'TIME', 'NAME', 'ACTION', 'ACTIVITIES', 'OUTCOMES', 'THUMBSUP', 'NUMLIKE']]
#        
#    
#        df_todo = pd.read_sql("select * from {} where NAME like '{}' and DATETIME like '{}'"\
#            .format(utils.db_todo, username, date), utils.ENGINE)
#        df_todo.columns = [x.upper() for x in df_todo.columns]
#        df_todo['NUMTODO'] = df_todo['TODO'].str.rstrip().str.split('\n').apply(len)
#        df_todo = df_todo[['DATETIME', 'TIME', 'NAME', 'OUTCOMES', 'TODO', 'NUMTODO']]
#    
#        #Note summary
#        note_duplicate, note_empty, note_nonsense, note_stat = summarize_tempo(df_note, 'NEXTACT', 5)
#            
#        #Action summary
#        action_duplicate, action_empty, action_nonsense, action_stat = summarize_tempo(df_action, 'ACTION', 10)
#        activity_duplicate, activity_empty, activity_nonsense, activity_stat = summarize_tempo(df_action, 'ACTIVITIES', 10)
#        outcome_duplicate, outcome_empty, outcome_nonsense, outcome_stat = summarize_tempo(df_action, 'OUTCOMES', 10)
#        
#        #Duplicate details
#        act_stat = pd.concat([action_stat, activity_stat, outcome_stat], axis = 1)
#        act_duplicate = pd.concat([action_duplicate, activity_duplicate, outcome_duplicate], axis = 1)
#    
#        writer = pd.ExcelWriter(
#            path + 'dailyreview_{}.xlsx'.format(datetime.now().strftime('%H%M%S'))
#        )
#        row_start = 0
#        col_start = 0
#        df_todo.to_excel(writer, sheet_name = username, startrow = row_start, startcol = col_start, index = False)
#        row_start += len(df_todo) + 2
#        df_action.to_excel(writer, sheet_name = username, startrow = row_start, startcol = col_start, index = False)
#        row_start += len(df_action) + 2
#        df_note.to_excel(writer, sheet_name = username, startrow = row_start, startcol = col_start, index = False)
#    
#        row_start = 0
#        col_start = 10
#        note_stat.to_excel(writer, sheet_name = username, startrow = row_start, startcol = col_start, index = False)
#        row_start += len(note_stat) + 2
#        note_duplicate.to_excel(writer, sheet_name = username, startrow = row_start, startcol = col_start, index = False)
#        row_start += len(note_duplicate) + 2
#        act_stat.to_excel(writer, sheet_name = username, startrow = row_start, startcol = col_start, index = False)
#        row_start += len(act_stat) + 2
#        act_duplicate.to_excel(writer, sheet_name = username, startrow = row_start, startcol = col_start, index = False)
#        
#        writer.save()
#    
#        return 'View in output directory'
##
#from logger import log_temponote, log_tempomonitor
#import traceback
#def get_weekly_report(username = utils.username, date = utils.current_date.strftime('%Y-%m-%d')):
#    level_result = utils.check_my_level(username)
#    if level_result == 0:
#        return "You are not allowed to check on this member"
#    else:
#        try:
#            date_input = datetime.strptime(date, '%Y-%m-%d')
#            delta = timedelta(days = datetime.now().weekday())
#            monday = date_input - delta
#            
#            date = datetime.strptime(date, '%Y-%m-%d').strftime('%d-%b-%y').upper()
#            path = utils.path_output + '{}_{}/'.format(date, username)
#            if not os.path.exists(path):
#                os.makedirs(path)
##            get_weekly_report(date='2020-10-05')
#            excelName = path + 'weeklyreview_{}.xlsx'.format(datetime.now().strftime('%H%M%S'))
#            writer = pd.ExcelWriter(excelName, engine = 'openpyxl')
#            firstData = []
#            firstData = pd.DataFrame(firstData)
#            firstData.to_excel(writer, sheet_name = "Monday", index = False)
#            writer.save()
#            book = load_workbook(excelName)
#            writer = pd.ExcelWriter(excelName, engine = 'openpyxl')
#            writer.book = book
#            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#            
#            for i in range(7):
#                date = (monday + timedelta(days = i)).strftime('%d-%b-%y').upper()
#                weekday = (monday + timedelta(days = i)).strftime("%A")
#                df_note = pd.read_sql("select * from {} where NAME like '{}' and DATETIME like '{}'"\
#                    .format(utils.db_temponote, username, date), utils.ENGINE)
#                df_note.columns = [x.upper() for x in df_note.columns]
#                df_note.sort_values(by='DATETIME', inplace=True, ascending=True)
#                df_note = df_note[['DATETIME', 'TIME', 'NAME', 'OUTCOME', 'NEXTACT']]
#            
#                df_action = pd.read_sql("select * from {} where NAME like '{}' and DATETIME like '{}'"\
#                    .format(utils.db_action, username, date), utils.ENGINE)
#                df_action.columns = [x.upper() for x in df_action.columns]
#                df_action.sort_values(by='DATETIME', inplace=True, ascending=True)
#                df_action['NUMLIKE'] = df_action['THUMBSUP'].str.split(', ').apply(len)
#                df_action = df_action[['DATETIME', 'TIME', 'NAME', 'ACTION', 'ACTIVITIES', 'OUTCOMES', 'THUMBSUP', 'NUMLIKE']]
#                
#                df_todo = pd.read_sql("select * from {} where NAME like '{}' and DATETIME like '{}'"\
#                    .format(utils.db_todo, username, date), utils.ENGINE)
#                df_todo.columns = [x.upper() for x in df_todo.columns]
#                df_todo['NUMTODO'] = df_todo['TODO'].str.rstrip().str.split('\n').apply(len)
#                df_todo = df_todo[['DATETIME', 'TIME', 'NAME', 'OUTCOMES', 'TODO', 'NUMTODO']]
#            
#                #Note summary
#                note_duplicate, note_empty, note_nonsense, note_stat = summarize_tempo(df_note, 'NEXTACT', 5)
#                    
#                #Action summary
#                action_duplicate, action_empty, action_nonsense, action_stat = summarize_tempo(df_action, 'ACTION', 10)
#                activity_duplicate, activity_empty, activity_nonsense, activity_stat = summarize_tempo(df_action, 'ACTIVITIES', 10)
#                outcome_duplicate, outcome_empty, outcome_nonsense, outcome_stat = summarize_tempo(df_action, 'OUTCOMES', 10)
#                
#                #Duplicate details
#                act_stat = pd.concat([action_stat, activity_stat, outcome_stat], axis = 1)
#                act_duplicate = pd.concat([action_duplicate, activity_duplicate, outcome_duplicate], axis = 1)
#    
#                row_start = 0
#                col_start = 0
#                df_todo.to_excel(writer, sheet_name = weekday, startrow = row_start, startcol = col_start, index = False)
#                row_start += len(df_todo) + 2
#                df_action.to_excel(writer, sheet_name = weekday, startrow = row_start, startcol = col_start, index = False)
#                row_start += len(df_action) + 2
#                df_note = df_note.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)
#                df_note.to_excel(writer, sheet_name = weekday, startrow = row_start, startcol = col_start, index = False)
#            
#                row_start = 0
#                col_start = 10
#                note_stat.to_excel(writer, sheet_name = weekday, startrow = row_start, startcol = col_start, index = False)
#                row_start += len(note_stat) + 2
#                note_duplicate.to_excel(writer, sheet_name = weekday, startrow = row_start, startcol = col_start, index = False)
#                row_start += len(note_duplicate) + 2
#                act_stat.to_excel(writer, sheet_name = weekday, startrow = row_start, startcol = col_start, index = False)
#                row_start += len(act_stat) + 2
#                act_duplicate.to_excel(writer, sheet_name = weekday, startrow = row_start, startcol = col_start, index = False)
##    encode("ascii",errors="ignore")
#            writer.save()
#            writer.close()
#        except:
#            log_tempomonitor.warning(traceback.format_exc())
#        return 'View in output directory'

def get_report(username = utils.username, date = utils.current_date.strftime('%Y-%m-%d'), mode = 'daily'):
    level_result = utils.check_my_level(username)
    if level_result == 0:
        return "You are not allowed to check on this member"
    else:
        if mode == 'daily':
            sheetName = username
            count = 1
        elif mode == 'weekly':
            date_input = datetime.strptime(date, '%Y-%m-%d')
            delta = timedelta(days = date_input.weekday())
            monday = date_input - delta
            sheetName = monday.strftime("%A")
            count = 7
        elif mode == 'monthly':
            month = datetime.strptime(date, '%Y-%m-%d').month
            year = datetime.strptime(date, '%Y-%m-%d').year
            day1 = datetime(year, month, 1)
            sheetName = day1.strftime("%d-%b")
            count = 32
        date = datetime.strptime(date, '%Y-%m-%d').strftime('%d-%b-%y').upper()
        path = utils.path_output + '{}_{}/'.format(date, username)
        if not os.path.exists(path):
            os.makedirs(path)
        
        if mode == 'daily':
            excelPath = path + 'dailyreview_{}.xlsx'.format(datetime.now().strftime('%H%M%S'))
        elif mode == 'weekly':
            excelPath = path + 'weeklyreview_{}.xlsx'.format(datetime.now().strftime('%H%M%S'))
        elif mode == 'monthly':
            excelPath = path + 'monthlyreview_{}.xlsx'.format(datetime.now().strftime('%H%M%S'))

        book = None
        try:
            book = load_workbook(excelPath)
        except Exception:
            log_temponote.warning('Creating new workbook')
        writer = pd.ExcelWriter(excelPath, engine = 'openpyxl')
        if book is not None:
            writer.book = book

        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        try:
            for i in range(count):
                if mode == 'weekly':
                    date = (monday + timedelta(days = i)).strftime('%d-%b-%y').upper()
                    sheetName = (monday + timedelta(days = i)).strftime("%A")
                elif mode == 'monthly':
                    if (day1 + timedelta(days = i)).month != month:
                        break
                    date = (day1 + timedelta(days = i)).strftime('%d-%b-%y').upper()
                    sheetName = (day1 + timedelta(days = i)).strftime("%d-%b")
                
                df_note = pd.read_sql("select * from {} where NAME like '{}' and DATETIME like '{}'"\
                    .format(utils.db_temponote, username, date), utils.ENGINE)
                df_note.columns = [x.upper() for x in df_note.columns]
                df_note.sort_values(by='DATETIME', inplace=True, ascending=True)
                df_note = df_note[['DATETIME', 'TIME', 'NAME', 'OUTCOME', 'NEXTACT']]
            
                df_action = pd.read_sql("select * from {} where NAME like '{}' and DATETIME like '{}'"\
                    .format(utils.db_action, username, date), utils.ENGINE)
                df_action.columns = [x.upper() for x in df_action.columns]
                df_action.sort_values(by='DATETIME', inplace=True, ascending=True)
                df_action['NUMLIKE'] = df_action['THUMBSUP'].str.split(', ').apply(len)
                df_action = df_action[['DATETIME', 'TIME', 'NAME', 'ACTION', 'ACTIVITIES', 'OUTCOMES', 'THUMBSUP', 'NUMLIKE']]
                
                df_todo = pd.read_sql("select * from {} where NAME like '{}' and DATETIME like '{}'"\
                    .format(utils.db_todo, username, date), utils.ENGINE)
                df_todo.columns = [x.upper() for x in df_todo.columns]
                df_todo['NUMTODO'] = df_todo['TODO'].str.rstrip().str.split('\n').apply(len)
                df_todo = df_todo[['DATETIME', 'TIME', 'NAME', 'OUTCOMES', 'TODO', 'NUMTODO']]

                #Note summary
                note_duplicate, note_empty, note_nonsense, note_stat = summarize_tempo(df_note, 'NEXTACT', 5)
                    
                #Action summary
                action_duplicate, action_empty, action_nonsense, action_stat = summarize_tempo(df_action, 'ACTION', 10)
                activity_duplicate, activity_empty, activity_nonsense, activity_stat = summarize_tempo(df_action, 'ACTIVITIES', 10)
                outcome_duplicate, outcome_empty, outcome_nonsense, outcome_stat = summarize_tempo(df_action, 'OUTCOMES', 10)
                
                #Duplicate details
                act_stat = pd.concat([action_stat, activity_stat, outcome_stat], axis = 1)
                act_duplicate = pd.concat([action_duplicate, activity_duplicate, outcome_duplicate], axis = 1)

                row_start = 0
                col_start = 0
                if i == 5:
                    df_recommendation = pd.read_sql("select * from {} where NAME like '{}' and DATETIME like '{}'"\
                        .format(utils.db_recommendation, username, date), utils.ENGINE)
                    df_recommendation.columns = [x.upper() for x in df_recommendation.columns]
                    df_recommendation['NUMRECOMMENDATION'] = df_recommendation['RECOMMENDATION'].str.rstrip().str.split('\n').apply(len)
                    df_recommendation = df_recommendation[['DATETIME', 'TIME', 'NAME', 'RECOMMENDATION','NUMRECOMMENDATION']]
                    df_recommendation.to_excel(writer, sheet_name = sheetName, startrow = row_start, startcol = col_start, index = False)
                    row_start += len(df_recommendation) + 2

                df_todo = df_todo.applymap(lambda x: ILLEGAL_CHARACTERS_RE.sub(r'', x) if isinstance(x, str) else x)
                df_todo.to_excel(writer, sheet_name = sheetName, startrow = row_start, startcol = col_start, index = False)
                row_start += len(df_todo) + 2
                df_action = df_action.applymap(lambda x: ILLEGAL_CHARACTERS_RE.sub(r'', x) if isinstance(x, str) else x)
                df_action.to_excel(writer, sheet_name = sheetName, startrow = row_start, startcol = col_start, index = False)
                row_start += len(df_action) + 2
                df_note = df_note.applymap(lambda x: ILLEGAL_CHARACTERS_RE.sub(r'', x) if isinstance(x, str) else x)
                df_note.to_excel(writer, sheet_name = sheetName, startrow = row_start, startcol = col_start, index = False)
            
                row_start = 0
                col_start = 10
                note_stat = note_stat.applymap(lambda x: ILLEGAL_CHARACTERS_RE.sub(r'', x) if isinstance(x, str) else x)
                note_stat.to_excel(writer, sheet_name = sheetName, startrow = row_start, startcol = col_start, index = False)
                row_start += len(note_stat) + 2
                note_duplicate = note_duplicate.applymap(lambda x: ILLEGAL_CHARACTERS_RE.sub(r'', x) if isinstance(x, str) else x)
                note_duplicate.to_excel(writer, sheet_name = sheetName, startrow = row_start, startcol = col_start, index = False)
                row_start += len(note_duplicate) + 2
                act_stat = act_stat.applymap(lambda x: ILLEGAL_CHARACTERS_RE.sub(r'', x) if isinstance(x, str) else x)
                act_stat.to_excel(writer, sheet_name = sheetName, startrow = row_start, startcol = col_start, index = False)
                row_start += len(act_stat) + 2
                act_duplicate = act_duplicate.applymap(lambda x: ILLEGAL_CHARACTERS_RE.sub(r'', x) if isinstance(x, str) else x)
                act_duplicate.to_excel(writer, sheet_name = sheetName, startrow = row_start, startcol = col_start, index = False)
        except Exception:
            log_temponote.warning(traceback.format_exc())

        writer.save()
        writer.close()
        return 'View in output directory'

def get_image(username = utils.username, date = utils.current_date.strftime('%Y-%m-%d')):
    level_result = utils.check_my_level(username)
    if level_result == 0:
        return "You are not allowed to check on this member"
    else:
        date = datetime.strptime(date, '%Y-%m-%d').strftime('%d-%b-%y').upper()
        path = utils.path_output + '{}_{}/'.format(date, username)
        if not os.path.exists(path):
            os.makedirs(path)
        
        df = pd.read_sql("select * from {} where NAME like '{}' and DATETIME like '{}'"\
                 .format(utils.db_tempomonitor, username, date), utils.ENGINE)
        df.columns = [x.upper() for x in df.columns]
        df.sort_values(by='DATETIME', inplace=True, ascending=True)
        #    if len(df) > num_image:
        #        df = df.sample(num_image).sort_values('DATETIME')
        #    for idx in df.index:
        #        frame = Image.open(io.BytesIO(base64.b64decode(df.loc[idx,'WEBCAM'])))
        #        frame.save(
        #            path + '{}_webcam.png'.format(df.loc[idx,'TIME'].replace(':', ''))
        #        )
        #        
        #        screen = Image.open(io.BytesIO(base64.b64decode(df.loc[idx,'SCREEN'])))
        #        screen.save(
        #            path + '{}_screen.png'.format(df.loc[idx,'TIME'].replace(':', ''))
        #        )
        date_position = (20,10)
        with imageio.get_writer( path +'webcams_screens.gif', mode='I', fps=5) as writer_screen:
            for idx, row in df.iterrows():
                screen = imageio.imread(base64.b64decode(row['SCREEN']))
                webcam = imageio.imread(base64.b64decode(row['WEBCAM']))
                concat_image = cv2.hconcat([webcam,screen])
                cv2.putText(
                 concat_image, #numpy array on which text is written
                 "{}".format(row['DATETIME']), #text
                 date_position, #position at which writing has to start
                 cv2.FONT_HERSHEY_SIMPLEX, #font family
                 0.4, #font size
                 (0,0,255)
                 )
                writer_screen.append_data(concat_image)
        writer_screen.close()
        
        return 'View in output directory'

def summarize_tempo(df, column_name, nonsense_value):
    try:
        df_duplicate = df.pivot_table(index = [column_name], aggfunc = 'size')
        df_duplicate = pd.DataFrame(df_duplicate[df_duplicate >= 2])
        df_duplicate.columns = ['{}_Duplicate'.format(column_name)]
        df_duplicate.reset_index(inplace = True)
    except:
        df_duplicate = pd.DataFrame(columns = [
            '{}'.format(column_name),     
            '{}_Duplicate'.format(column_name)
        ])
    
    df[column_name] = df[column_name].str.strip()
    df_empty = df[(df[column_name].isna()) | (df[column_name] == '')]

    df_nonsense = df[df[column_name].str.len() <= nonsense_value]
    
    df_stat = pd.DataFrame()
    df_stat['{}_Duplicate'.format(column_name)] = [len(df_duplicate)]
    df_stat['{}_Empty'.format(column_name)] = len(df_empty)
    df_stat['{}_NonSense'.format(column_name)] = len(df_nonsense)
    
    return df_duplicate, df_empty, df_nonsense, df_stat


def get_time_use(name = None):   
#    trường hợp 1
    if name is not None:
        df = pd.read_sql("SELECT qt_temponote.name,qt_memberinfo.code,qt_temponote.datetime \
        FROM {} INNER JOIN {} on to_char(qt_memberinfo.code) = to_char(qt_temponote.name)\
        where DATETIME in (select min(DATETIME) from {} where name like '{}')"\
        .format(utils.db_memberinfo,utils.db_temponote,utils.db_temponote,name),utils.ENGINE)
        
 #    trường hợp 2
    if name is None:
        df = pd.read_sql("SELECT SUBQUERY.NAME code,{}.NAME name,\
        (SELECT min(DATETIME) FROM {} WHERE NAME LIKE SUBQUERY.NAME )DATETIME\
        FROM (SELECT DISTINCT to_char(NAME)NAME FROM {}) SUBQUERY\
        join {} on to_char({}.code) = SUBQUERY.NAME"\
        .format(utils.db_memberinfo,utils.db_temponote,utils.db_temponote,utils.db_memberinfo,utils.db_memberinfo),utils.ENGINE)
    
    df.to_excel('the-first-user-MAL.xlsx',index=False)
    os.system("start EXCEL.EXE the-first-user-MAL.xlsx")
               
    return 'Done'
def get_time_working(username = None,start = utils.current_date.strftime('%Y-%m-%d')):
    
    path = utils.path_output + '{}_{}/'.format(start, username)
    if not os.path.exists(path):
        os.makedirs(path)
     #    trường hợp 1       
    if username is not None:
        df = pd.read_sql("SELECT * from {} where name like '{}' \
                 and (LOWER(nextact) like LOWER('checkout') or LOWER(nextact) like LOWER('checkin')) \
             and datetime like TO_DATE ('{}','YYYY-MM-DD') ORDER BY to_char(name), datetime ASC"\
             .format(utils.db_temponote,username,start),utils.ENGINE)
    #    trường hợp 2
#    if username is None:
#        df = pd.read_sql("SELECT * from {} where (LOWER(nextact) like \
#             LOWER('checkout') or LOWER(nextact) like LOWER('checkin')) \
#             and datetime BETWEEN TO_DATE ('{}','YYYY-MM-DD')\
#             and TO_DATE ('{}','YYYY-MM-DD') ORDER BY to_char(name), datetime ASC"\
#             .format(utils.db_temponote,start,end),utils.ENGINE)
        
    excelName = path + 'working_time_{}.xlsx'.format(datetime.now().strftime('%H%M%S'))
    writer = pd.ExcelWriter(excelName)
    df.to_excel(writer, sheet_name = 'All username', index = False)
    return 'View in output directory'
        